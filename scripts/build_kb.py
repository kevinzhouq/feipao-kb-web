import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Missing openpyxl. Use the bundled Codex Python runtime or install openpyxl.") from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT.parent / "知识库" / "客服知识库.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "kb.json"
SHEET_NAME = "知识库"


def clean(value):
    if value is None:
        return ""
    text = str(value).replace("\u2028", "\n").replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"[ \t]+", " ", text).strip()


def main():
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not source.exists():
        raise SystemExit(f"Knowledge base file not found: {source}")

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(cell) for cell in next(rows)]
    required = ["类型", "问题", "答复", "后续处理", "备注"]
    indexes = {name: headers.index(name) for name in required if name in headers}
    missing = [name for name in required if name not in indexes]
    if missing:
        raise SystemExit(f"Missing columns: {', '.join(missing)}")

    records = []
    for row_number, row in enumerate(rows, start=2):
        record = {name: clean(row[indexes[name]] if indexes[name] < len(row) else "") for name in required}
        if not record["问题"] and not record["答复"]:
            continue
        records.append({
            "id": f"kb-{row_number}",
            "type": record["类型"] or "未分类",
            "question": record["问题"],
            "answer": record["答复"],
            "followUp": record["后续处理"],
            "note": record["备注"],
        })

    payload = {
        "source": str(source),
        "sheet": SHEET_NAME,
        "recordCount": len(records),
        "categories": sorted({item["type"] for item in records}),
        "records": records,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(records)} records to {output}")


if __name__ == "__main__":
    main()
